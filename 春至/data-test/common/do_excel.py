import time
from openpyxl import load_workbook
from common import read_path
from common.read_config import ReadConfig
from common.do_mysql import DoMySql


class DoExcel:
    def __init__(self, file_name):
        self.file_name = file_name
        self.wb = load_workbook(self.file_name)

    def get_title(self, sheet_name):
        """获取每一列数据的标题"""
        sheet = self.wb[sheet_name]
        title = []
        for i in range(1, sheet.max_column + 1):
            title.append(sheet.cell(1, i).value)
        return title

    def get_test_data(self, sheet_name):
        """获取测试数据"""
        sheet = self.wb[sheet_name]
        test_data = []
        title = self.get_title(sheet_name)

        # 测试运行的模式
        mode = ReadConfig().read_config(read_path.conf_path, 'MODE', 'mode')
        # 测试运行的用例编号列表
        case_id_list = ReadConfig().read_config(read_path.conf_path, 'MODE', 'case_id_list')
        # 获取环境flag
        env_flag = ReadConfig().read_config(read_path.conf_path, 'MODE', 'env_flag')
        # 获取接口域名
        if env_flag == '0':  # 如果是0，表示测试环境
            domain_name = ReadConfig().read_config(read_path.conf_path, 'TEST', 'domain_name')
        else:  # 如果是1，表示生产环境
            domain_name = ReadConfig().read_config(read_path.conf_path, 'PRODUCTION', 'domain_name')
        # 按行读取，从第2行开始
        for i in range(2, sheet.max_row + 1):
            row_data = {}
            # 按列读取，从第1列开始
            for j in range(1, sheet.max_column + 1):
                row_data[title[j - 1]] = sheet.cell(i, j).value
            row_data['Url'] = domain_name + row_data['Url']

            test_data.append(row_data)
        # 如果mode=1，执行所有测试用例
        if mode == '1':
            final_data = test_data
        # 如果mode !=1，执行case_id_list里面的用例
        else:
            final_data = []
            for item in test_data:
                if item['Case_Id'] in eval(case_id_list):
                    final_data.append(item)
        return final_data

    def write_back(self, row, col, new_value, sheet_name):
        """将测试结果写回测试用例文件"""
        sheet = self.wb[sheet_name]
        sheet.cell(row, col).value = new_value
        self.wb.save(self.file_name)

    def add_time_stamp(self, data):
        """给请求参数添加时间戳"""
        for i in data:
            param_dic = eval(i['Param'])
            param_dic['time_stamp'] = int(time.time())
            i['Param'] = str(param_dic)
        return data

    def get_base_data(self, sheet_name):
        """获取配置数据"""
        sheet = self.wb[sheet_name]
        base_data = []
        title = self.get_title(sheet_name)
        # 按行读取，从第2行开始
        for i in range(2, sheet.max_row + 1):
            row_data = {}
            # 按列读取，从第1列开始
            for j in range(1, sheet.max_column + 1):
                row_data[title[j - 1]] = sheet.cell(i, j).value
            base_data.append(row_data)
        return base_data

    def clear_test_data(self, sheet_name):
        """清除测试数据"""
        sheet = self.wb[sheet_name]
        # 按行读取，从第2行开始
        for i in range(2, sheet.max_row + 1):
            self.write_back(i, 8, '', sheet_name)   # HttpCheck
            self.write_back(i, 9, '', sheet_name)   # CountCheck
            self.write_back(i, 10, '', sheet_name)  # DataCheck
            self.write_back(i, 11, '', sheet_name)  # FailReason
            self.write_back(i, 12, '', sheet_name)  # TestResult


if __name__ == '__main__':
    # data = DoExcel(read_path.test_data_path).get_base_data('country_code')
    # data2 = DoExcel(read_path.test_data_path).get_test_data('test_data')
    # print(data2)
    DoExcel(read_path.test_data_path).clear_test_data('test_data')
