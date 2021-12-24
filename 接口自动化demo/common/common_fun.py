# from baseView.baseView import BaseView
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
import logging
import time,os
import csv
import datetime
from xlrd import open_workbook
from openpyxl import load_workbook


class Common(object):

    # 获取脚本执行时间
    def getTime(self):
        self.now = time.strftime("%Y-%m-%d %H_%M_%S")
        return self.now

    # # 截屏
    # def getScreenShot(self, module):
    #     time = self.getTime()
    #     image_file = os.path.dirname(os.path.dirname(__file__))+'/screenshots/%s_%s.png' % (module, time)
    #     logging.info('get %s screenshot' % module)
    #     self.driver.get_screenshot_as_file(image_file)

    # 读取csv文件数据
    def get_csv_data(self, csv_file, line):
        logging.info('========get_scv_data=======')
        with open(csv_file, 'r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            for index, row in enumerate(reader, 1):  # enumerate(a,start) a是可迭代对象，start是计数起始数字
                if index == line:
                    return row

    # 获取今天多少号，和以后几天多少号
    def start_end_Time(self, day):
        startDay = datetime.date.today()
        endDay = datetime.timedelta(days=day)
        tomorrow = (startDay + endDay).strftime("%d")
        today = startDay.strftime("%d")
        return startDay, endDay

    def getExcelTestData(self, excel_path):
        '''
            这种的获取方式，Excel的空，读取出来是 空 不是 null
        :param excel_path:
        :return:
        '''
        # 打开 excel 文件，获取数据列表
        openExcelFile = load_workbook(excel_path)
        # 读取第一 sheet 页的数据
        getSheet = openExcelFile.active
        # 获取工作表
        dataList = []
        # 数据List
        for row in getSheet.rows:
            list1 = []
            for cell in row:
                if cell.value == None:
                    cell.value = ""
                list1.append(cell.value)
            dataList.append(list1)
        # print(dataList)
        dataList.pop(0)
        # print(dataList)
        return dataList

    def getExcelTestData_sheetName(self, excel_path, Sheet_name):
        '''
            这种的获取方式，Excel的空，读取出来是 空 不是 null
        :param excel_path:
        :return:
        '''
        # 文件必须是xlsx格式，如果是其他格式在执行前可利用win32辅助转化
        # 打开 excel 文件，获取数据列表
        openExcelFile = load_workbook(excel_path)
        # 获取某一特定的工作表
        sheet = openExcelFile.get_sheet_by_name(Sheet_name)
        # 获取数据列表
        dataList = []
        # 数据List
        for row in sheet.rows:
            list1 = []
            for cell in row:
                if cell.value == None:
                    cell.value = ""
                list1.append(cell.value)
            dataList.append(list1)
        # print(dataList)
        dataList.pop(0)
        # print(dataList)
        return dataList


if __name__ == '__main__':
    c = Common()
    from common import read_path
    xlsx_path = read_path.test_data_path
    # excelPath = r'E:\python项目\LHB\location_PC\data\requireements_excedValues\场地需求之活动信息.xlsx'
    # c.getExcelTestData(excelPath)
    c.getExcelTestData_sheetName(xlsx_path,'杭州')

