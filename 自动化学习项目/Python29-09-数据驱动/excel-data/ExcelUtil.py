from openpyxl import load_workbook


class ParseExcel(object):
    def __init__(self, excelPath, sheetName):
        # 读取excel文件
        self.wb = load_workbook(excelPath)
        # 通过工作表名称获取一个工作表对象
        self.sheet = self.wb.get_sheet_by_name(sheetName)
        # 获取工作表中存在数据的区域的最大行号
        self.maxRowNum = self.sheet.max_row

    def getDatasFromSheet(self):
        # 用于存放从工作表中读取出来的数据
        dataList = []
        # 因为工作表中的第一行是标题行，所以需要去掉
        for line in self.sheet.rows:
            # 1.遍历工作表中数据区域的每一行
            # 2.将每行中各个单元格的数据取出存于列表tmpList中
            # 3.再将存放一行数据的列表添加到最络数据列表dataList中
            tmpList = []
            tmpList.append(line[1].value)
            print("line[1].value:", line[1].value)
            tmpList.append(line[2].value)
            print("line[2].value:", line[2].value)
            dataList.append(tmpList)
        print(dataList[1:])
        return dataList[1:]


if __name__ == '__main__':
    excelPath = r'D:\Python项目\自动化学习项目\Python29-09-数据驱动\excel-data\测试数据.xlsx'
    sheetName = "搜索数据表"
    pe = ParseExcel(excelPath, sheetName)
    print(pe.getDatasFromSheet())
    for i in pe.getDatasFromSheet():
        print(i[0], i[1])
